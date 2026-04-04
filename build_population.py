import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load ──────────────────────────────────────────────────────────────────────
BASE          = Path(__file__).parent
compiled_path = BASE / "Fulfillment_Compilation.xlsx"
output_path   = BASE / "Population_For_Calculation.xlsx"

print("Loading compiled data...")
df = pd.read_excel(compiled_path)
print(f"  Loaded: {len(df):,} rows, {df['FOLIO'].nunique():,} unique properties")

# ── Type coercions ────────────────────────────────────────────────────────────
df['LIKELY DEAL SCORE']   = pd.to_numeric(df['LIKELY DEAL SCORE'],   errors='coerce')
df['SCORE']               = pd.to_numeric(df['SCORE'],               errors='coerce')
df['MARKETING DM COUNT']  = pd.to_numeric(df['MARKETING DM COUNT'],  errors='coerce').fillna(0)
df['MARKETING SMS COUNT'] = pd.to_numeric(df['MARKETING SMS COUNT'], errors='coerce').fillna(0)
df['VACANT_num']          = pd.to_numeric(df['VACANT'],              errors='coerce').fillna(0)
df['Month_dt']            = pd.to_datetime(df['Month'], format='%Y-%m')

month_abbr = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
              7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
df['Month_label'] = df['Month_dt'].apply(
    lambda d: f"{month_abbr[d.month]}-{str(d.year)[2:]}"
)

# ── Action Plans -> numeric days ──────────────────────────────────────────────
def extract_days(val):
    if pd.isna(val): return np.nan
    m = re.match(r'(\d+)\s*DAY', str(val).upper())
    return int(m.group(1)) if m else np.nan

df['ACTION_DAYS'] = df['ACTION PLANS'].apply(extract_days)

# ── Distress (long form) ──────────────────────────────────────────────────────
distress_main_cols = ['MAIN DISTRESS #1','MAIN DISTRESS #2','MAIN DISTRESS #3','MAIN DISTRESS #4']
no_distress_re = re.compile(r'^No distress', re.IGNORECASE)

dist_long = (
    df[['FOLIO','Month_label','Source_File'] + distress_main_cols]
    .melt(id_vars=['FOLIO','Month_label','Source_File'],
          value_vars=distress_main_cols, value_name='dtype')
    .dropna(subset=['dtype'])
)
dist_long = dist_long[~dist_long['dtype'].str.match(no_distress_re, na=True)].copy()
dist_long['dtype'] = dist_long['dtype'].str.strip()

# Add Vacant from binary column
vacant_rows = df[df['VACANT_num'] > 0][['FOLIO','Month_label','Source_File']].copy()
vacant_rows['dtype'] = 'Vacant'
dist_long = pd.concat([dist_long, vacant_rows], ignore_index=True)

# Deduplicate: same FOLIO + month + source + dtype = 1 occurrence
dist_deduped = dist_long.drop_duplicates(subset=['FOLIO','Month_label','Source_File','dtype'])

distress_list = (
    dist_deduped.groupby('FOLIO')['dtype']
    .apply(lambda s: ', '.join(sorted(s.unique())))
    .rename('Distress_List')
)
distress_count = (
    dist_deduped.groupby('FOLIO')['dtype']
    .nunique()
    .rename('Distress_Count')
)

# ── Marketing count: latest month per FOLIO ───────────────────────────────────
df['total_mkt'] = df['MARKETING DM COUNT'] + df['MARKETING SMS COUNT']
latest_by_folio = df.groupby('FOLIO')['Month_dt'].max().rename('latest_month')
df_j = df.join(latest_by_folio, on='FOLIO')
df_latest = df_j[df_j['Month_dt'] == df_j['latest_month']]

mkt_agg = df_latest.groupby('FOLIO').agg(
    Latest_Marketing_Count=('total_mkt',           'sum'),
    DM_Count_Latest        =('MARKETING DM COUNT', 'sum'),
    SMS_Count_Latest       =('MARKETING SMS COUNT','sum'),
)

# ── Scores & Action Plan ──────────────────────────────────────────────────────
def most_frequent_days(series):
    s = series.dropna()
    if s.empty: return np.nan
    counts = s.value_counts()
    tied = counts[counts == counts.max()].index.tolist()
    return int(min(tied))

agg_scores = df.groupby('FOLIO').agg(
    Max_Likelihood_Score=('LIKELY DEAL SCORE','max'),
    Avg_Likelihood_Score=('LIKELY DEAL SCORE','mean'),
    Max_Total_Score     =('SCORE','max'),
    Avg_Total_Score     =('SCORE','mean'),
    Max_Action_Plan     =('ACTION_DAYS','min'),
)
avg_action = (
    df.groupby('FOLIO')['ACTION_DAYS']
    .apply(most_frequent_days)
    .rename('Avg_Action_Plan')
)

# ── Months appeared & appearances ─────────────────────────────────────────────
months_appeared = (
    df.groupby('FOLIO')['Month_label']
    .apply(lambda s: ', '.join(sorted(s.unique(),
           key=lambda x: pd.to_datetime(x, format='%b-%y'))))
    .rename('Months_Appeared')
)
appearances_count = df.groupby('FOLIO').size().rename('Appearances_Count')

# ── Source files ──────────────────────────────────────────────────────────────
source_files = (
    df.groupby('FOLIO')['Source_File']
    .apply(lambda s: ' | '.join(sorted(s.unique())))
    .rename('Source_Files')
)

# ── Address ───────────────────────────────────────────────────────────────────
address = df.groupby('FOLIO')['ADDRESS'].first().rename('Address')

# ── Assemble ──────────────────────────────────────────────────────────────────
print("Assembling population table...")
pop = (
    agg_scores
    .join(avg_action)
    .join(mkt_agg)
    .join(distress_list)
    .join(distress_count)
    .join(months_appeared)
    .join(appearances_count)
    .join(source_files)
    .join(address)
    .reset_index()
)

pop['Avg_Likelihood_Score'] = pop['Avg_Likelihood_Score'].round(1)
pop['Avg_Total_Score']      = pop['Avg_Total_Score'].round(1)

day_label = {30:'30 day', 60:'60 day', 90:'90 day'}
pop['Max_Action_Plan'] = pop['Max_Action_Plan'].map(day_label)
pop['Avg_Action_Plan'] = pop['Avg_Action_Plan'].map(day_label)

pop['Distress_List']  = pop['Distress_List'].fillna('None')
pop['Distress_Count'] = pop['Distress_Count'].fillna(0).astype(int)

col_order = [
    'FOLIO','Address',
    'Months_Appeared','Appearances_Count',
    'Max_Likelihood_Score','Avg_Likelihood_Score',
    'Max_Total_Score','Avg_Total_Score',
    'Max_Action_Plan','Avg_Action_Plan',
    'DM_Count_Latest','SMS_Count_Latest','Latest_Marketing_Count',
    'Distress_List','Distress_Count',
    'Source_Files',
]
pop = pop[col_order]
print(f"  Rows: {len(pop):,}  |  Columns: {len(pop.columns)}")

# ── Validation ────────────────────────────────────────────────────────────────
likely_bins   = [-0.5, 20, 40, 60, 80, 100.5]
likely_labels = ['20-0','40-21','60-41','80-61','100-81']
check = pd.cut(pop['Max_Likelihood_Score'], bins=likely_bins, labels=likely_labels).value_counts().sort_index()
print("\nLikely Deal Score distribution (MAX):")
print(check.to_string())
print("\nAction Plan (MAX):")
print(pop['Max_Action_Plan'].value_counts().sort_index().to_string())
print(f"\nTotal distress occurrences: {pop['Distress_Count'].sum():,}")

# ── Write with pandas (fast), then style header only ─────────────────────────
print(f"\nWriting Excel (fast mode)...")
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    pop.to_excel(writer, index=False, sheet_name='Population')

print("Applying header styling...")
wb = load_workbook(output_path)
ws = wb['Population']

thin_side = Side(style='thin', color='CCCCCC')
thin = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

HEADER_FONT  = Font(name='Helvetica', bold=True, size=10, color='FFFFFF')
HEADER_FILL  = PatternFill('solid', fgColor='0B5394')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border    = thin
ws.row_dimensions[1].height = 36

# Column widths
col_widths = {
    'FOLIO':18,'Address':30,
    'Months_Appeared':22,'Appearances_Count':14,
    'Max_Likelihood_Score':20,'Avg_Likelihood_Score':20,
    'Max_Total_Score':18,'Avg_Total_Score':18,
    'Max_Action_Plan':16,'Avg_Action_Plan':16,
    'DM_Count_Latest':16,'SMS_Count_Latest':16,'Latest_Marketing_Count':22,
    'Distress_List':50,'Distress_Count':16,
    'Source_Files':70,
}
for i, col_name in enumerate(pop.columns, start=1):
    ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 16)

ws.freeze_panes = 'C2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(pop.columns))}1"

wb.save(output_path)
print(f"Saved: {output_path}")
print("Done.")
