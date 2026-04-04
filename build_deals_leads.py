"""
build_deals_leads.py
─────────────────────────────────────────────────────────────────────────────
Generates:
  1. Documents/Deals and Leads Atlas - Enriched.xlsx
       • Adds a 'Data_Source' column to each property:
           'Fulfillment'          → data pulled from Fulfillment_Compilation (MAX logic)
           'Deals and Leads Atlas'→ data pulled from the deals/leads file itself (fallback)
  2. Expected_Result_Max.xlsx  (overwrites previous version)
       • Column C  — Properties in the fulfillment   (all 37k properties, MAX logic)
       • Column F  — Clients Lead                    (848 leads, best available data)
       • Column G  — Client Deals                    (151 deals, best available data)

Data resolution priority (same for both F and G):
  PRIMARY  → Fulfillment_Compilation.xlsx joined on FOLIO (MAX across all months)
  FALLBACK → Deals and Leads Atlas.xlsx own values
             (used when FOLIO is missing or not present in fulfillment)
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE            = Path(__file__).parent
COMPILED_PATH   = BASE / "Fulfillment_Compilation.xlsx"
DEALS_PATH      = BASE / "Documents" / "Deals and Leads Atlas.xlsx"
OUTPUT_MAX         = BASE / "Expected_Result_Max_FG.xlsx"
OUTPUT_MAX_FF_ONLY = BASE / "Expected_Result_Max_FG_InFulfillment.xlsx"
OUTPUT_ENRICHED     = BASE / "Documents" / "Deals and Leads Atlas - Enriched.xlsx"
OUTPUT_FULFILLMENT  = BASE / "Documents" / "Deals and Leads Atlas - In Fulfillment.xlsx"

# ── Shared bucket definitions ─────────────────────────────────────────────────
likely_bins   = [-0.5, 20, 40, 60, 80, 100.5]
likely_labels = ['20-0', '40-21', '60-41', '80-61', '100-81']
score_bins    = [-0.5, 200, 400, 600, 800, 1000.5]
score_labels  = ['200-0', '400-201', '600-401', '800-601', '1000-801']
mkt_bins      = [0.5, 5, 10, 19, 1e9]
mkt_labels    = ['1 to 5', '6 to 10', '11 to 19', '20+']
action_map    = {30: '30 day', 60: '60 day', 90: '90 day'}

# ── Column order in the report ────────────────────────────────────────────────
COLS = [
    'Category',                          # A (index 0)
    'Total Properties',                  # B (index 1)
    'Properties in the fulfillment',     # C (index 2)
    'Sold',                              # D (index 3)
    'Market Deals',                      # E (index 4)
    'Clients Lead',                      # F (index 5)
    'Client Deals',                      # G (index 6)
    'Sold Concentration %Sold',          # H (index 7)
    'Sold to investors concentration',   # I (index 8)
    'Client Deals Concentration',        # J (index 9)
    'Client Leads Concentration',        # K (index 10)
]
SECTION_NAMES = {'Likely Deal Score', 'Total Score', 'Action Plan', 'Mkt Count', 'Distress'}

# ── Helpers ───────────────────────────────────────────────────────────────────
def extract_days(val):
    if pd.isna(val): return np.nan
    m = re.match(r'(\d+)\s*DAY', str(val).upper())
    return int(m.group(1)) if m else np.nan

def count_by_bin(series, bins, labels):
    bucketed = pd.cut(series, bins=bins, labels=labels)
    counts = bucketed.value_counts()
    return {lbl: int(counts.get(lbl, 0)) for lbl in labels}

def count_by_action(series, mapping):
    mapped = series.map(mapping)
    counts = mapped.value_counts()
    return {c: int(counts.get(c, 0)) for c in ['30 day', '60 day', '90 day']}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: Load fulfillment → MAX aggregations per FOLIO
# ══════════════════════════════════════════════════════════════════════════════
print("Loading fulfillment data...")
df = pd.read_excel(COMPILED_PATH)
print(f"  {len(df):,} rows, {df['FOLIO'].nunique():,} unique FOLIOs")

df['LIKELY DEAL SCORE']   = pd.to_numeric(df['LIKELY DEAL SCORE'],   errors='coerce')
df['SCORE']               = pd.to_numeric(df['SCORE'],               errors='coerce')
df['MARKETING DM COUNT']  = pd.to_numeric(df['MARKETING DM COUNT'],  errors='coerce').fillna(0)
df['MARKETING SMS COUNT'] = pd.to_numeric(df['MARKETING SMS COUNT'], errors='coerce').fillna(0)
df['Month_dt']            = pd.to_datetime(df['Month'], format='%Y-%m')
df['ACTION_DAYS']         = df['ACTION PLANS'].apply(extract_days)

# Latest-month DM+SMS per FOLIO
df['total_mkt'] = df['MARKETING DM COUNT'] + df['MARKETING SMS COUNT']
latest_by_folio = df.groupby('FOLIO')['Month_dt'].max().rename('latest_month')
df_j   = df.join(latest_by_folio, on='FOLIO')
ff_mkt = (df_j[df_j['Month_dt'] == df_j['latest_month']]
          .groupby('FOLIO')['total_mkt'].sum()
          .rename('ff_mkt'))

# First marketing recommendation date per FOLIO (earliest month in fulfillment)
ff_first_date = df.groupby('FOLIO')['Month_dt'].min().rename('ff_first_date')

# MAX scores and best action plan per FOLIO
ff_agg = df.groupby('FOLIO').agg(
    ff_likely = ('LIKELY DEAL SCORE', 'max'),
    ff_score  = ('SCORE',             'max'),
    ff_days   = ('ACTION_DAYS',       'min'),   # min days = best plan (30 < 60 < 90)
).join(ff_mkt, how='left').join(ff_first_date, how='left')

# Distress from fulfillment (MAIN DISTRESS #1-4 columns + binary VACANT)
distress_main_cols = ['MAIN DISTRESS #1', 'MAIN DISTRESS #2', 'MAIN DISTRESS #3', 'MAIN DISTRESS #4']
no_dist_re = re.compile(r'^No distress', re.IGNORECASE)

dist_long = (
    df[['FOLIO'] + distress_main_cols]
    .melt(id_vars='FOLIO', value_vars=distress_main_cols, value_name='dtype')
    .dropna(subset=['dtype'])
)
dist_long = dist_long[~dist_long['dtype'].str.match(no_dist_re, na=True)].copy()
dist_long['dtype'] = dist_long['dtype'].str.strip()

vac_num   = pd.to_numeric(df['VACANT'], errors='coerce').fillna(0)
vac_rows  = pd.DataFrame({'FOLIO': df[vac_num > 0]['FOLIO'].unique(), 'dtype': 'Vacant'})
dist_long = pd.concat([dist_long, vac_rows], ignore_index=True)

# Build FOLIO → set-of-distress-types lookup
ff_distress_map = dist_long.groupby('FOLIO')['dtype'].apply(set).to_dict()

print(f"  Aggregation ready for {len(ff_agg):,} FOLIOs")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: Column C — counts for all fulfillment properties (MAX logic)
# ══════════════════════════════════════════════════════════════════════════════
distress_order = [
    'Pre-foreclosure', 'Vacant', 'Taxes (Tax Delinquent)', 'Estate (Pre-Probate)',
    'Inter family transfer', 'Probate', 'Judgement',
    'Liens city/county', 'Liens other', 'Liens Mechanic',
    'Default Risk', 'High Equity', '55+ (Senior)',
    'Absentee', 'Absentee Out of State', 'Downsizing',
]
for t in sorted(dist_long['dtype'].unique()):
    if t not in distress_order:
        distress_order.append(t)

C_dist = {dtype: int(dist_long[dist_long['dtype'] == dtype]['FOLIO'].nunique())
          for dtype in distress_order}

C_likely = count_by_bin(ff_agg['ff_likely'], likely_bins, likely_labels)
C_score  = count_by_bin(ff_agg['ff_score'],  score_bins,  score_labels)
C_action = count_by_action(ff_agg['ff_days'], action_map)
C_mkt    = count_by_bin(ff_agg['ff_mkt'],    mkt_bins,   mkt_labels)

print("\n=== COLUMN C (fulfillment MAX) ===")
print("Likely Deal Score:", C_likely)
print("Action Plan:", C_action)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3: Load Deals and Leads Atlas + enrich with fulfillment data
# ══════════════════════════════════════════════════════════════════════════════
print("\nLoading Deals and Leads Atlas...")
dl = pd.read_excel(DEALS_PATH)
n_leads = (dl['PROPERTY STATUS'] == 'Lead').sum()
n_deals = (dl['PROPERTY STATUS'] == 'Deal').sum()
print(f"  {len(dl):,} rows — {n_leads} Leads / {n_deals} Deals")

dl['LIKELY DEAL SCORE']   = pd.to_numeric(dl['LIKELY DEAL SCORE'],   errors='coerce')
dl['SCORE']               = pd.to_numeric(dl['SCORE'],               errors='coerce')
dl['MARKETING DM COUNT']  = pd.to_numeric(dl['MARKETING DM COUNT'],  errors='coerce').fillna(0)
dl['MARKETING SMS COUNT'] = pd.to_numeric(dl['MARKETING SMS COUNT'], errors='coerce').fillna(0)
dl['dl_days'] = dl['ACTION PLANS'].apply(extract_days)
dl['dl_mkt']  = dl['MARKETING DM COUNT'] + dl['MARKETING SMS COUNT']

# Normalize FOLIO key for join (null FOLIOs won't match anything → fallback)
dl['FOLIO_key'] = dl['FOLIO'].astype(str).str.strip()
dl.loc[dl['FOLIO_key'] == 'nan', 'FOLIO_key'] = np.nan

# Join fulfillment MAX aggregations
ff_for_join = ff_agg.reset_index().rename(columns={'FOLIO': 'FOLIO_key'})
dl = dl.merge(ff_for_join, on='FOLIO_key', how='left')

# Binary distress column → distress-order label mapping (for fallback properties)
BINARY_DIST_MAP = {
    'PRE-FORECLOSURE':       'Pre-foreclosure',
    'VACANT':                'Vacant',
    'TAXES':                 'Taxes (Tax Delinquent)',
    'ESTATE':                'Estate (Pre-Probate)',
    'INTER FAMILY TRANSFER': 'Inter family transfer',
    'PROBATE':               'Probate',
    'JUDGEMENT':             'Judgement',
    'LIENS CITY/COUNTY':     'Liens city/county',
    'LIENS OTHER':           'Liens other',
    'LIENS MECHANIC':        'Liens Mechanic',
    'HIGH EQUITY':           'High Equity',
    '55+':                   '55+ (Senior)',
    'ABSENTEE':              'Absentee',
    'DOWNSIZING':            'Downsizing',
}

def get_binary_distress(row):
    """Read individual binary distress columns from the deals/leads file."""
    types = set()
    for col, label in BINARY_DIST_MAP.items():
        if col in row.index:
            val = pd.to_numeric(row[col], errors='coerce')
            if not pd.isna(val) and val > 0:
                types.add(label)
    return types

# Resolve each row: use fulfillment data if available, else use file's own data
print("  Resolving data sources (primary: Fulfillment | fallback: Deals and Leads Atlas)...")

def resolve_row(row):
    in_ff = pd.notna(row.get('ff_likely'))
    if in_ff:
        return pd.Series({
            'Data_Source':          'Fulfillment',
            'r_likely':             row['ff_likely'],
            'r_score':              row['ff_score'],
            'r_days':               row['ff_days'],
            'r_mkt':                max(row['ff_mkt'], 1) if pd.notna(row.get('ff_mkt')) else 1,
            'r_distress':           ff_distress_map.get(str(row['FOLIO_key']).strip(), set()),
            'First_Marketing_Date': row.get('ff_first_date'),
        })
    else:
        return pd.Series({
            'Data_Source':          'Deals and Leads Atlas',
            'r_likely':             row['LIKELY DEAL SCORE'],
            'r_score':              row['SCORE'],
            'r_days':               row['dl_days'],
            'r_mkt':                row['dl_mkt'],
            'r_distress':           get_binary_distress(row),
            'First_Marketing_Date': pd.NaT,
        })

resolved = dl.apply(resolve_row, axis=1)
dl = pd.concat([dl, resolved], axis=1)

n_ff  = (dl['Data_Source'] == 'Fulfillment').sum()
n_fb  = (dl['Data_Source'] == 'Deals and Leads Atlas').sum()
print(f"  Data source -> Fulfillment: {n_ff} | Deals and Leads Atlas (fallback): {n_fb}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4: Save enriched Deals and Leads Atlas
# Adds a 'Data_Source' column right after 'PROPERTY STATUS'
# ══════════════════════════════════════════════════════════════════════════════
print(f"\nSaving enriched Deals and Leads Atlas...")

# Columns to exclude from enriched output (working columns added during processing)
_working_cols = {
    'FOLIO_key', 'ff_likely', 'ff_score', 'ff_days', 'ff_mkt', 'ff_first_date',
    'dl_days', 'dl_mkt', 'r_likely', 'r_score', 'r_days', 'r_mkt', 'r_distress',
    'Data_Source', 'First_Marketing_Date',   # re-inserted at correct positions below
}
orig_cols = [c for c in dl.columns if c not in _working_cols]

# Insert Data_Source immediately after PROPERTY STATUS
try:
    ps_idx = orig_cols.index('PROPERTY STATUS') + 1
except ValueError:
    ps_idx = len(orig_cols)
out_cols = orig_cols[:ps_idx] + ['Data_Source', 'First_Marketing_Date'] + orig_cols[ps_idx:]

dl[out_cols].to_excel(OUTPUT_ENRICHED, index=False, sheet_name='Sheet1')
print(f"  Saved: {OUTPUT_ENRICHED}")

# ── Fulfillment-only file (349 matched properties) ────────────────────────────
print(f"\nSaving Deals and Leads Atlas - In Fulfillment...")
dl_ff = dl[dl['Data_Source'] == 'Fulfillment'][out_cols].copy()
n_ff_leads = (dl_ff['PROPERTY STATUS'] == 'Lead').sum()
n_ff_deals = (dl_ff['PROPERTY STATUS'] == 'Deal').sum()
dl_ff.to_excel(OUTPUT_FULFILLMENT, index=False, sheet_name='Sheet1')
print(f"  {len(dl_ff):,} rows — {n_ff_leads} Leads / {n_ff_deals} Deals")
print(f"  Saved: {OUTPUT_FULFILLMENT}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5: Column F (Leads) and Column G (Deals) bucket counts
# ══════════════════════════════════════════════════════════════════════════════
print("\nComputing Leads and Deals bucket counts...")

def compute_col_counts(sub_df):
    """Compute all report-section bucket counts for a leads or deals subset."""
    likely_d = count_by_bin(sub_df['r_likely'], likely_bins, likely_labels)
    score_d  = count_by_bin(sub_df['r_score'],  score_bins,  score_labels)
    action_d = count_by_action(sub_df['r_days'], action_map)
    mkt_d    = count_by_bin(sub_df['r_mkt'],    mkt_bins,   mkt_labels)

    # Distress: count unique rows where the distress type appears (set membership)
    dist_d = {}
    for dtype in distress_order:
        dist_d[dtype] = int(
            sub_df['r_distress'].apply(
                lambda s: dtype in s if isinstance(s, set) else False
            ).sum()
        )
    return likely_d, score_d, action_d, mkt_d, dist_d

leads_df = dl[dl['PROPERTY STATUS'] == 'Lead'].copy()
deals_df = dl[dl['PROPERTY STATUS'] == 'Deal'].copy()

F_likely, F_score, F_action, F_mkt, F_dist = compute_col_counts(leads_df)
G_likely, G_score, G_action, G_mkt, G_dist = compute_col_counts(deals_df)

print("\n=== COLUMN F (Leads) ===")
print("Likely Deal Score:", F_likely)
print("Action Plan:", F_action)

print("\n=== COLUMN G (Deals) ===")
print("Likely Deal Score:", G_likely)
print("Action Plan:", G_action)

# Fulfillment-only subsets (345 Leads + 4 Deals)
leads_ff_df = leads_df[leads_df['Data_Source'] == 'Fulfillment'].copy()
deals_ff_df = deals_df[deals_df['Data_Source'] == 'Fulfillment'].copy()

FF_likely, FF_score, FF_action, FF_mkt, FF_dist = compute_col_counts(leads_ff_df)
GG_likely, GG_score, GG_action, GG_mkt, GG_dist = compute_col_counts(deals_ff_df)

print("\n=== COLUMN F — Fulfillment only (Leads) ===")
print("Likely Deal Score:", FF_likely)
print("Action Plan:", FF_action)

print("\n=== COLUMN G — Fulfillment only (Deals) ===")
print("Likely Deal Score:", GG_likely)
print("Action Plan:", GG_action)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6: Write Expected_Result_Max.xlsx  (columns C, F, G populated)
# ══════════════════════════════════════════════════════════════════════════════
def build_rows(C_lik, C_sco, C_act, C_mkt, C_dis,
               F_lik, F_sco, F_act, F_mkt, F_dis,
               G_lik, G_sco, G_act, G_mkt, G_dis):

    n = len(COLS)

    def section_header(name):
        r = list(COLS); r[0] = name; return r

    def data_row(label, c, f, g):
        r = [''] * n
        r[0] = label
        r[2] = c    # Column C — Properties in the fulfillment
        r[5] = f    # Column F — Clients Lead
        r[6] = g    # Column G — Client Deals
        return r

    def blank():
        return [''] * n

    rows = []

    # Likely Deal Score
    rows.append(section_header('Likely Deal Score'))
    for lbl in ['100-81', '80-61', '60-41', '40-21', '20-0']:
        rows.append(data_row(lbl, C_lik[lbl], F_lik[lbl], G_lik[lbl]))
    rows.append(data_row('Total', sum(C_lik.values()), sum(F_lik.values()), sum(G_lik.values())))
    rows.append(blank())

    # Total Score
    rows.append(section_header('Total Score'))
    for lbl in ['1000-801', '800-601', '600-401', '400-201', '200-0']:
        rows.append(data_row(lbl, C_sco[lbl], F_sco[lbl], G_sco[lbl]))
    rows.append(data_row('Total', sum(C_sco.values()), sum(F_sco.values()), sum(G_sco.values())))
    rows.append(blank())

    # Action Plan
    rows.append(section_header('Action Plan'))
    for lbl in ['30 day', '60 day', '90 day']:
        rows.append(data_row(lbl, C_act[lbl], F_act[lbl], G_act[lbl]))
    rows.append(data_row('Total', sum(C_act.values()), sum(F_act.values()), sum(G_act.values())))
    rows.append(blank())

    # Mkt Count
    rows.append(section_header('Mkt Count'))
    for lbl in ['1 to 5', '6 to 10', '11 to 19', '20+']:
        rows.append(data_row(lbl, C_mkt[lbl], F_mkt[lbl], G_mkt[lbl]))
    rows.append(data_row('Total', sum(C_mkt.values()), sum(F_mkt.values()), sum(G_mkt.values())))
    rows.append(blank())

    # Distress
    rows.append(section_header('Distress'))
    for dtype in distress_order:
        rows.append(data_row(dtype, C_dis[dtype], F_dis[dtype], G_dis[dtype]))
    rows.append(data_row('Total', sum(C_dis.values()), sum(F_dis.values()), sum(G_dis.values())))

    return rows


def write_excel(path, rows, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font   = Font(name='Helvetica', bold=True, size=10, color='FFFFFF')
    header_fill   = PatternFill('solid', fgColor='0B5394')
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    section_font  = Font(name='Georgia', bold=True, size=10)
    section_fill  = PatternFill('solid', fgColor='D9E2F3')
    section_align = Alignment(horizontal='left', vertical='center')
    total_font    = Font(name='Helvetica', bold=True, size=10)
    total_fill    = PatternFill('solid', fgColor='F2F2F2')
    data_font     = Font(name='Helvetica', size=10)
    count_font    = Font(name='Helvetica', bold=True, size=10, color='0B5394')
    center_align  = Alignment(horizontal='center', vertical='center')
    left_align    = Alignment(horizontal='left',   vertical='center')
    thin_side     = Side(style='thin', color='CCCCCC')
    thin          = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

    # Header row
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = thin
    ws.row_dimensions[1].height = 36

    # Data rows
    COUNT_COLS = [2, 5, 6]   # C, F, G (0-based)
    for row_data in rows:
        ws.append(row_data)
        row_idx = ws.max_row
        label   = row_data[0]

        if label in SECTION_NAMES:
            for cell in ws[row_idx]:
                cell.font = section_font; cell.fill = section_fill
                cell.alignment = section_align; cell.border = thin
        elif label == 'Total':
            for cell in ws[row_idx]:
                cell.font = total_font; cell.fill = total_fill; cell.border = thin
            ws[row_idx][0].alignment = left_align
            for ci in COUNT_COLS:
                ws[row_idx][ci].alignment = center_align
        elif label == '':
            pass
        else:
            for cell in ws[row_idx]:
                cell.font = data_font; cell.alignment = left_align; cell.border = thin
            for ci in COUNT_COLS:
                ws[row_idx][ci].font = count_font
                ws[row_idx][ci].alignment = center_align

    ws.column_dimensions['A'].width = 32
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 22
    ws.freeze_panes = 'A2'

    wb.save(path)
    print(f"  Saved: {path}")


rows = build_rows(
    C_likely, C_score, C_action, C_mkt, C_dist,
    F_likely, F_score, F_action, F_mkt, F_dist,
    G_likely, G_score, G_action, G_mkt, G_dist,
)

print(f"\nWriting Expected_Result_Max_FG.xlsx...")
write_excel(OUTPUT_MAX, rows, "Max Logic")

rows_ff = build_rows(
    C_likely, C_score, C_action, C_mkt, C_dist,
    FF_likely, FF_score, FF_action, FF_mkt, FF_dist,
    GG_likely, GG_score, GG_action, GG_mkt, GG_dist,
)

print(f"Writing Expected_Result_Max_FG_InFulfillment.xlsx...")
write_excel(OUTPUT_MAX_FF_ONLY, rows_ff, "Max Logic - In Fulfillment")

print("\nAll done.")
