import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Load ──────────────────────────────────────────────────────────────────────
BASE          = Path(__file__).parent
compiled_path = BASE / "Fulfillment_Compilation.xlsx"
print("Loading compiled data...")
df = pd.read_excel(compiled_path)
print(f"  Loaded: {len(df):,} rows")

# ── Type coercions ─────────────────────────────────────────────────────────────
df['LIKELY DEAL SCORE'] = pd.to_numeric(df['LIKELY DEAL SCORE'], errors='coerce')
df['SCORE']             = pd.to_numeric(df['SCORE'], errors='coerce')
df['MARKETING DM COUNT']  = pd.to_numeric(df['MARKETING DM COUNT'],  errors='coerce').fillna(0)
df['MARKETING SMS COUNT'] = pd.to_numeric(df['MARKETING SMS COUNT'], errors='coerce').fillna(0)
df['Month_dt'] = pd.to_datetime(df['Month'], format='%Y-%m')

# ── Action Plans: extract numeric days ───────────────────────────────────────
def extract_days(val):
    if pd.isna(val):
        return np.nan
    m = re.match(r'(\d+)\s*DAY', str(val).upper())
    return int(m.group(1)) if m else np.nan

df['ACTION_DAYS'] = df['ACTION PLANS'].apply(extract_days)

# ── Distress (shared for both files) ─────────────────────────────────────────
distress_main_cols = ['MAIN DISTRESS #1', 'MAIN DISTRESS #2', 'MAIN DISTRESS #3', 'MAIN DISTRESS #4']
no_distress_re = re.compile(r'^No distress', re.IGNORECASE)

dist_long = (
    df[['FOLIO'] + distress_main_cols]
    .melt(id_vars='FOLIO', value_vars=distress_main_cols, value_name='dtype')
    .dropna(subset=['dtype'])
)
dist_long = dist_long[~dist_long['dtype'].str.match(no_distress_re, na=True)].copy()
dist_long['dtype'] = dist_long['dtype'].str.strip()

# Ordered distress list (matches template, then any extras)
distress_order = [
    'Pre-foreclosure', 'Vacant', 'Taxes (Tax Delinquent)', 'Estate (Pre-Probate)',
    'Inter family transfer', 'Probate', 'Judgement',
    'Liens city/county', 'Liens other', 'Liens Mechanic',
    'Default Risk', 'High Equity', '55+ (Senior)',
    'Absentee', 'Absentee Out of State', 'Downsizing',
]
for t in sorted(dist_long['dtype'].unique()):
    if t not in distress_order:
        distress_order.append(t)

# Count unique FOLIOs per distress type
distress_counts = {}
for dtype in distress_order:
    folios_with = dist_long[dist_long['dtype'] == dtype]['FOLIO'].nunique()
    distress_counts[dtype] = int(folios_with)

# Special case: "Vacant" does not appear in MAIN DISTRESS columns.
# Use the binary VACANT column instead.
df_vacant_check = df[['FOLIO', 'VACANT']].copy()
df_vacant_check['VACANT_num'] = pd.to_numeric(df_vacant_check['VACANT'], errors='coerce')
vacant_folio_count = df_vacant_check[df_vacant_check['VACANT_num'] > 0]['FOLIO'].nunique()
distress_counts['Vacant'] = int(vacant_folio_count)
print(f"  Vacant (from binary column): {vacant_folio_count:,}")

print(f"\nDistress types found: {len(distress_order)}")
for d in distress_order:
    print(f"  {d}: {distress_counts[d]:,}")

# ── Marketing count: latest month per FOLIO, sum DM + SMS ─────────────────────
df['total_mkt'] = df['MARKETING DM COUNT'] + df['MARKETING SMS COUNT']
latest_month_by_folio = df.groupby('FOLIO')['Month_dt'].max().rename('latest_month')
df_joined = df.join(latest_month_by_folio, on='FOLIO')
df_latest = df_joined[df_joined['Month_dt'] == df_joined['latest_month']]
mkt_count = df_latest.groupby('FOLIO')['total_mkt'].sum().rename('mkt_count')

# ── MAX per FOLIO ─────────────────────────────────────────────────────────────
print("\nComputing MAX aggregation...")
agg_max = df.groupby('FOLIO').agg(
    max_likely=('LIKELY DEAL SCORE', 'max'),
    max_score=('SCORE', 'max'),
    min_days=('ACTION_DAYS', 'min'),   # min days = best category (30 < 60 < 90)
).join(mkt_count, how='left')

total_properties = len(agg_max)
print(f"  Unique properties: {total_properties:,}")

# ── AVERAGE per FOLIO ─────────────────────────────────────────────────────────
print("Computing AVERAGE aggregation...")

def most_frequent_days(series):
    s = series.dropna()
    if s.empty:
        return np.nan
    counts = s.value_counts()
    max_count = counts.max()
    tied = counts[counts == max_count].index.tolist()
    return min(tied)   # tie -> lowest days = highest priority (30 > 60 > 90)

avg_score = df.groupby('FOLIO').agg(
    avg_likely=('LIKELY DEAL SCORE', 'mean'),
    avg_score=('SCORE', 'mean'),
).join(mkt_count, how='left')

modal_days = df.groupby('FOLIO')['ACTION_DAYS'].apply(most_frequent_days).rename('modal_days')
agg_avg = avg_score.join(modal_days, how='left')

# ── Binning helpers ───────────────────────────────────────────────────────────
likely_bins   = [-0.5, 20, 40, 60, 80, 100.5]
likely_labels = ['20-0', '40-21', '60-41', '80-61', '100-81']

score_bins    = [-0.5, 200, 400, 600, 800, 1000.5]
score_labels  = ['200-0', '400-201', '600-401', '800-601', '1000-801']

mkt_bins      = [0.5, 5, 10, 19, 1e9]
mkt_labels    = ['1 to 5', '6 to 10', '11 to 19', '20+']

action_map = {30: '30 day', 60: '60 day', 90: '90 day'}


def count_by_bin(series, bins, labels):
    bucketed = pd.cut(series, bins=bins, labels=labels)
    counts = bucketed.value_counts()
    return {lbl: int(counts.get(lbl, 0)) for lbl in labels}


def count_by_action(series, mapping):
    mapped = series.map(mapping)
    counts = mapped.value_counts()
    cats = ['30 day', '60 day', '90 day']
    return {c: int(counts.get(c, 0)) for c in cats}


# ── Build counts ──────────────────────────────────────────────────────────────
print("Building MAX counts...")
mx_likely = count_by_bin(agg_max['max_likely'], likely_bins, likely_labels)
mx_score  = count_by_bin(agg_max['max_score'],  score_bins,  score_labels)
mx_action = count_by_action(agg_max['min_days'], action_map)
mx_mkt    = count_by_bin(agg_max['mkt_count'],  mkt_bins,   mkt_labels)

print("Building AVERAGE counts...")
av_likely = count_by_bin(agg_avg['avg_likely'],   likely_bins, likely_labels)
av_score  = count_by_bin(agg_avg['avg_score'],    score_bins,  score_labels)
av_action = count_by_action(agg_avg['modal_days'], action_map)
av_mkt    = count_by_bin(agg_avg['mkt_count'],    mkt_bins,   mkt_labels)

print("\n=== MAX SUMMARY ===")
print("Likely Deal Score:", mx_likely)
print("Total Score:", mx_score)
print("Action Plan:", mx_action)
print("Mkt Count:", mx_mkt)

print("\n=== AVERAGE SUMMARY ===")
print("Likely Deal Score:", av_likely)
print("Total Score:", av_score)
print("Action Plan:", av_action)
print("Mkt Count:", av_mkt)

# ── Column definitions ────────────────────────────────────────────────────────
COLS = [
    'Category',
    'Total Properties',
    'Properties in the fulfillment',
    'Sold',
    'Market Deals',
    'Clients Lead',
    'Client Deals',
    'Sold Concentration %Sold',
    'Sold to investors concentration',
    'Client Deals Concentration',
    'Client Leads Concentration',
]

SECTION_NAMES = {
    'Likely Deal Score', 'Total Score', 'Action Plan', 'Mkt Count', 'Distress'
}


def build_rows(likely_d, score_d, action_d, mkt_d):
    rows = []
    n = len(COLS)

    def section_header(name):
        r = list(COLS)
        r[0] = name
        return r

    def data_row(label, count):
        r = [''] * n
        r[0] = label
        r[2] = count
        return r

    def blank():
        return [''] * n

    # Likely Deal Score
    rows.append(section_header('Likely Deal Score'))
    for lbl in ['100-81', '80-61', '60-41', '40-21', '20-0']:
        rows.append(data_row(lbl, likely_d[lbl]))
    rows.append(data_row('Total', sum(likely_d.values())))
    rows.append(blank())

    # Total Score
    rows.append(section_header('Total Score'))
    for lbl in ['1000-801', '800-601', '600-401', '400-201', '200-0']:
        rows.append(data_row(lbl, score_d[lbl]))
    rows.append(data_row('Total', sum(score_d.values())))
    rows.append(blank())

    # Action Plan
    rows.append(section_header('Action Plan'))
    for lbl in ['30 day', '60 day', '90 day']:
        rows.append(data_row(lbl, action_d[lbl]))
    rows.append(data_row('Total', sum(action_d.values())))
    rows.append(blank())

    # Mkt Count
    rows.append(section_header('Mkt Count'))
    for lbl in ['1 to 5', '6 to 10', '11 to 19', '20+']:
        rows.append(data_row(lbl, mkt_d[lbl]))
    rows.append(data_row('Total', sum(mkt_d.values())))
    rows.append(blank())

    # Distress
    rows.append(section_header('Distress'))
    for dtype in distress_order:
        rows.append(data_row(dtype, distress_counts[dtype]))
    rows.append(data_row('Total', sum(distress_counts.values())))

    return rows


def write_excel(path, rows, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Styles
    header_font  = Font(name='Helvetica', bold=True, size=10, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='0B5394')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    section_font = Font(name='Georgia', bold=True, size=10)
    section_fill = PatternFill('solid', fgColor='D9E2F3')
    section_align = Alignment(horizontal='left', vertical='center')

    total_font   = Font(name='Helvetica', bold=True, size=10)
    total_fill   = PatternFill('solid', fgColor='F2F2F2')

    data_font    = Font(name='Helvetica', size=10)
    count_font   = Font(name='Helvetica', bold=True, size=10, color='0B5394')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left', vertical='center')

    thin_side = Side(style='thin', color='CCCCCC')
    thin = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

    # Write header row
    ws.append(COLS)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin
    ws.row_dimensions[1].height = 36

    # Write data rows
    for row_data in rows:
        ws.append(row_data)
        row_idx = ws.max_row
        label = row_data[0]

        if label in SECTION_NAMES:
            for cell in ws[row_idx]:
                cell.font      = section_font
                cell.fill      = section_fill
                cell.alignment = section_align
                cell.border    = thin
        elif label == 'Total':
            for cell in ws[row_idx]:
                cell.font   = total_font
                cell.fill   = total_fill
                cell.border = thin
            ws[row_idx][0].alignment = left_align
            ws[row_idx][2].alignment = center_align
        elif label == '':
            pass   # blank separator row
        else:
            for cell in ws[row_idx]:
                cell.font      = data_font
                cell.alignment = left_align
                cell.border    = thin
            # Count cell (col C)
            c = ws[row_idx][2]
            c.font      = count_font
            c.alignment = center_align

    # Column widths
    ws.column_dimensions['A'].width = 32
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 22
    ws.freeze_panes = 'A2'

    wb.save(path)
    print(f"  Saved: {path}")


print("\nWriting Expected_Result_Max.xlsx...")
write_excel(BASE / "Expected_Result_Max.xlsx", build_rows(mx_likely, mx_score, mx_action, mx_mkt), "Max Logic")

print("Writing Expected_Result_Average.xlsx...")
write_excel(BASE / "Expected_Result_Average.xlsx", build_rows(av_likely, av_score, av_action, av_mkt), "Average Logic")

print("\nAll done.")
