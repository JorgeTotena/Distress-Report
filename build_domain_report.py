"""
build_domain_report.py
─────────────────────────────────────────────────────────────────────────────
Generates Expected_Result_Max_BCDFG_InFulfillment.xlsx with:
  Column B — Total Properties        (all domain properties with BUYBOX SCORE > 0)
  Column C — Properties in the fulfillment (MAX logic across the actual fulfillment window)
  Column D — Sold since start of fulfillment window (pre-sale scores from fulfillment)
  Column E — Market Deals            (overlap: market deals ∩ fulfillment-sold)
  Column F — Clients Lead            (fulfillment-matched leads/appointments/dead leads/contracts)
  Column G — Client Deals            (fulfillment-matched deals only)
  Column H — Sold Concentration %              =D/C  (formula)
  Column I — Sold to Investors Concentration   =E/C  (formula)
  Column J — Client Deals Conc.               =G/C  (formula)
  Column K — Client Leads Conc.               =F/C  (formula)

Window: defined entirely by the .xlsx files present in Fulfillments/. SOLD_SINCE / WINDOW_START
is the first day of the earliest fulfillment month; WINDOW_END is the first day of the latest.
The user controls the window by adding or removing files from the folder.
Deals/Leads: any .xlsx in Documents/ whose name contains "leads" or "deals" (case-insensitive).
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

BASE              = Path(__file__).parent
FULFILLMENTS_DIR  = BASE / "Fulfillments"
COMPILED_PATH     = BASE / "Fulfillment_Compilation.xlsx"   # written each run for audit
DOCS_DIR          = BASE / "Documents"
DOMAIN_DIR        = BASE / "Domain Full Data"
MARKET_PATH       = BASE / "Market Deals" / "Market Deals Pillar Home Buyers.xlsx"
CLIENT_NAME       = "Pillar Home Buyers"
OUTPUT            = BASE / f"{CLIENT_NAME} - Distress Report - {pd.Timestamp.today().strftime('%Y-%m')}.xlsx"
# DR_PATH removed — DEFAULT RISK is now a column in the domain file

# ── Client start date (used to filter leads and deals) ───────────────────────
# Pillar Home Buyers started doing business with 8020REI on ~2024-05-24
CLIENT_START_DATE = pd.to_datetime("2024-05-24")
print(f"Client since {CLIENT_START_DATE.date()} (informational only — leads/deals are filtered to the fulfillment window start, computed below)")

# ── Analysis window ──────────────────────────────────────────────────────────
# The window is defined entirely by the files present in Fulfillments/.
# WINDOW_START / WINDOW_END / SOLD_SINCE are derived from the earliest and
# latest fulfillment-file months in STEP 1 below. The user controls the
# window by adding or removing files from the folder.

# ── Bucket definitions ────────────────────────────────────────────────────────
likely_bins   = [-0.5, 20, 40, 60, 80, 100.5]
likely_labels = ['20-0', '40-21', '60-41', '80-61', '100-81']
score_bins    = [-0.5, 200, 400, 600, 800, 1000.5]
score_labels  = ['200-0', '400-201', '600-401', '800-601', '1000-801']
mkt_bins      = [0.5, 5, 10, 19, 1e9]
mkt_labels    = ['1 to 5', '6 to 10', '11 to 19', '20+']
action_map    = {30: '30 day', 60: '60 day', 90: '90 day'}

distress_order = [
    'Pre-foreclosure', 'Vacant', 'Taxes (Tax Delinquent)', 'Estate (Pre-Probate)',
    'Inter family transfer', 'Probate', 'Judgement',
    'Liens city/county', 'Liens other', 'Liens Mechanic',
    'Default Risk', 'High Equity', '55+ (Senior)',
    'Absentee', 'Absentee Out of State', 'Downsizing',
    'Owner Occupied',
]

DOMAIN_DIST_MAP = {
    'PRE-FORECLOSURE':       'Pre-foreclosure',
    'VACANT':                'Vacant',
    'TAXES':                 'Taxes (Tax Delinquent)',
    'ESTATE':                'Estate (Pre-Probate)',
    'INTER FAMILY TRANSFER': 'Inter family transfer',
    'PROBATE':               'Probate',
    'JUDGEMENT':             'Judgement',
    'LIENS CITY/COUNTY':     'Liens city/county',
    'LIENS OTHER':           'Liens other',
    'LIENS MECHANIC':        'Liens Mechanic',
    'DEFAULT RISK':          'Default Risk',
    'HIGH EQUITY':           'High Equity',
    '55+':                   '55+ (Senior)',
    # ABSENTEE handled separately: 1=in-state, 2=out-of-state, 0=owner occupied
    'DOWNSIZING':            'Downsizing',
}

COLS = [
    'Category', 'Total Properties', 'Properties in the fulfillment',
    'Sold', 'Market Deals', 'Clients Lead', 'Client Deals',
    'Sold Concentration %Sold', 'Sold to investors concentration',
    'Client Deals Concentration', 'Client Leads Concentration',
]
ZIP_SECTION_TITLE = 'ZIP Code — most Client Deals first (BUYBOX SCORE > 0)'
SECTION_NAMES = {'Likely Deal Score', 'Total Score', 'Action Plan', 'Mkt Count', 'Distress', ZIP_SECTION_TITLE}

# ── Default Risk lookup — derived from domain (DEFAULT RISK column) ───────────
# Loaded early so dr_folios is available for fulfillment injection (STEP 1)
# Some domain file versions omit this column — handled gracefully.
print("Loading Default Risk from domain...")
_parquet_early = DOMAIN_DIR / "domain.parquet"
dr_folios = set()
if _parquet_early.exists():
    import pyarrow.parquet as _pq
    if 'DEFAULT RISK' in _pq.read_schema(_parquet_early).names:
        _dr = pd.read_parquet(_parquet_early, columns=['PROPERTY ID (BUYBOX)', 'FOLIO', 'DEFAULT RISK'])
        _dr_risk = _dr[pd.to_numeric(_dr['DEFAULT RISK'], errors='coerce').fillna(0) > 0]
        dr_folios = set(_dr_risk['FOLIO'].dropna().astype(str).str.strip().unique())
        del _dr, _dr_risk
    else:
        print("  DEFAULT RISK column not in domain file — Default Risk counts will be 0")
else:
    _xlsx = sorted(f for f in DOMAIN_DIR.glob("*.xlsx") if not f.name.startswith("~$"))
    _sample = pd.read_excel(_xlsx[0], nrows=0)
    if 'DEFAULT RISK' in _sample.columns:
        _dr = pd.concat([pd.read_excel(f, usecols=['PROPERTY ID (BUYBOX)', 'FOLIO', 'DEFAULT RISK']) for f in _xlsx], ignore_index=True)
        _dr_risk = _dr[pd.to_numeric(_dr['DEFAULT RISK'], errors='coerce').fillna(0) > 0]
        dr_folios = set(_dr_risk['FOLIO'].dropna().astype(str).str.strip().unique())
        del _dr, _dr_risk
    else:
        print("  DEFAULT RISK column not in domain xlsx files — Default Risk counts will be 0")
print(f"  {len(dr_folios):,} unique FOLIOs with Default Risk")

# ── Helpers ───────────────────────────────────────────────────────────────────
def extract_days(val):
    if pd.isna(val): return np.nan
    m = re.match(r'(\d+)\s*DAY', str(val).upper())
    return int(m.group(1)) if m else np.nan

def count_by_bin(series, bins, labels):
    bucketed = pd.cut(series.dropna(), bins=bins, labels=labels)
    counts = bucketed.value_counts()
    return {lbl: int(counts.get(lbl, 0)) for lbl in labels}

def count_by_action(series, mapping):
    mapped = series.dropna().map(mapping)
    counts = mapped.value_counts()
    return {c: int(counts.get(c, 0)) for c in ['30 day', '60 day', '90 day']}

def domain_distress_counts(df_sub):
    result = {dtype: 0 for dtype in distress_order}
    for dom_col, label in DOMAIN_DIST_MAP.items():
        if dom_col in df_sub.columns:
            vals = pd.to_numeric(df_sub[dom_col], errors='coerce').fillna(0)
            result[label] = int((vals > 0).sum())
    # ABSENTEE: 1 = in-state, 2 = out-of-state, 0 = owner occupied
    # Do NOT fillna — NaN means unknown, not owner-occupied
    if 'ABSENTEE' in df_sub.columns:
        abs_vals = pd.to_numeric(df_sub['ABSENTEE'], errors='coerce')
        result['Absentee']              = int((abs_vals == 1).sum())
        result['Absentee Out of State'] = int((abs_vals == 2).sum())
        result['Owner Occupied']        = int((abs_vals == 0).sum())
    # DEFAULT RISK is now a direct column in the domain file — handled by DOMAIN_DIST_MAP loop above
    return result


def fulfillment_distress_counts(df_sub, dist_long_lookup):
    """Distress counts using fulfillment binary flags as source of truth.

    The fulfillment file's binary distress columns (PRE-FORECLOSURE > 0,
    TAXES > 0, ESTATE > 0, …) are authoritative for any column compared
    against fulfillment (C/D/E/F/G). Domain often has fewer binary flags
    set per FOLIO, so reading from domain there undercounts.

    ABSENTEE (3-way: Absentee / Absentee Out of State / Owner Occupied) is also
    fulfillment-sourced — via the FOLIO-keyed ff_abs map (max across months) — so
    it reflects the point-in-time status at recommendation, not domain's current-
    only value. See absentee_counts_from_ff / ff_abs.
    """
    folios_set = set(df_sub['FOLIO'].dropna().astype(str).str.strip().unique())
    result = {dtype: 0 for dtype in distress_order}
    if folios_set:
        _key = dist_long_lookup['FOLIO'].astype(str).str.strip()
        _sub = dist_long_lookup[_key.isin(folios_set)]
        for dtype in distress_order:
            result[dtype] = int(_sub[_sub['dtype'] == dtype]['FOLIO'].nunique())
    result.update(absentee_counts_from_ff(df_sub['FOLIO']))
    return result


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: Compile fulfillment files → MAX aggregations per FOLIO
# Window = whatever .xlsx files are in Fulfillments/. The user controls the
# window by adding or removing files. SOLD_SINCE = first day of the earliest
# file's month.
# ══════════════════════════════════════════════════════════════════════════════
print("Compiling fulfillment files...")
_ff_files = sorted(f for f in FULFILLMENTS_DIR.glob("*.xlsx") if not f.name.startswith("~$"))
if not _ff_files:
    raise FileNotFoundError(f"No fulfillment files found in {FULFILLMENTS_DIR}")

_ff_months   = sorted({f.name[:7] for f in _ff_files})
WINDOW_START = pd.Timestamp(_ff_months[0])
WINDOW_END   = pd.Timestamp(_ff_months[-1])
SOLD_SINCE   = WINDOW_START   # Column D cutoff — first day of the earliest fulfillment month
print(f"  Window: {WINDOW_START.strftime('%Y-%m')} – {WINDOW_END.strftime('%Y-%m')} "
      f"({len(_ff_months)} month(s), {len(_ff_files)} file(s))")

_parts = []
for _f in _ff_files:
    _part = pd.read_excel(_f)
    _part['Month']              = _f.name[:7]   # YYYY-MM from filename prefix
    _part['Marketing_Channel']  = 'DM' if 'Direct Mail' in _f.name else ('SMS' if 'SMS' in _f.name else 'Other')
    _part['Source_File']        = _f.name
    _parts.append(_part)
    print(f"  {_f.name}: {len(_part):,} rows")

df = pd.concat(_parts, ignore_index=True)
print(f"  Total: {len(df):,} rows, {df['FOLIO'].nunique():,} unique FOLIOs")
# Compilation file is written at the end of STEP 5, after enrichment with
# LAST SALE DATE, PROPERTY STATUS, and MARKET DEAL flag for validation.

df['LIKELY DEAL SCORE']   = pd.to_numeric(df['LIKELY DEAL SCORE'],   errors='coerce')
df['SCORE']               = pd.to_numeric(df['SCORE'],               errors='coerce')
for _col in ['MARKETING DM COUNT', 'MARKETING SMS COUNT']:
    df[_col] = pd.to_numeric(df[_col], errors='coerce').fillna(0) if _col in df.columns else 0
df['Month_dt']            = pd.to_datetime(df['Month'], format='%Y-%m')
df['ACTION_DAYS']         = df['ACTION PLANS'].apply(extract_days)
df['total_mkt']           = df['MARKETING DM COUNT'] + df['MARKETING SMS COUNT']

# Marketing count: max of (total DM across all months, total SMS across all months)
# e.g. 7 SMS + 4 DM → 7, not 11 — measures strongest channel, not combined
ff_dm  = df.groupby('FOLIO')['MARKETING DM COUNT'].sum()
ff_sms = df.groupby('FOLIO')['MARKETING SMS COUNT'].sum()
ff_mkt = pd.concat([ff_dm, ff_sms], axis=1).max(axis=1).rename('ff_mkt')

ff_agg = df.groupby('FOLIO').agg(
    ff_likely = ('LIKELY DEAL SCORE', 'max'),
    ff_score  = ('SCORE',             'max'),
    ff_days   = ('ACTION_DAYS',       'min'),
).join(ff_mkt, how='left')

# Source of truth for binary distress flags = fulfillment binary columns
# (PRE-FORECLOSURE > 0, VACANT > 0, TAXES > 0, …) per FOLIO across all months.
# Replaces the old MAIN DISTRESS #1-4 ranking, which silently dropped flags
# that didn't make the top-4 ranked slot for a given property.
ff_binary_cols = [c for c in DOMAIN_DIST_MAP.keys() if c in df.columns]
_long_parts = []
for _col in ff_binary_cols:
    _vals = pd.to_numeric(df[_col], errors='coerce').fillna(0)
    _flagged = df.loc[_vals > 0, 'FOLIO'].dropna().unique()
    if len(_flagged):
        _long_parts.append(pd.DataFrame({'FOLIO': _flagged, 'dtype': DOMAIN_DIST_MAP[_col]}))
dist_long = (pd.concat(_long_parts, ignore_index=True)
             if _long_parts else pd.DataFrame(columns=['FOLIO', 'dtype']))
# Inject Default Risk for fulfillment FOLIOs that appear in the DR file
ff_folios_set  = set(df['FOLIO'].dropna().astype(str).str.strip().unique())
dr_in_ff_folios = ff_folios_set & dr_folios
dr_ff_rows = pd.DataFrame({'FOLIO': list(dr_in_ff_folios), 'dtype': 'Default Risk'})
dist_long = pd.concat([dist_long, dr_ff_rows], ignore_index=True)
ff_distress_map = dist_long.groupby('FOLIO')['dtype'].apply(set).to_dict()

# ── ABSENTEE per FOLIO from the FULFILLMENT file (max across months) ──────────
# ABSENTEE is a 3-way code (0 = owner-occupied, 1 = in-state absentee,
# 2 = out-of-state). The fulfillment file is the source of truth for C/D/E/F/G
# (domain holds only current ownership, not the point-in-time status when the
# property was recommended). MAX across the window = 'most-distressed status ever
# seen' (2 > 1 > 0), consistent with the 'ever-flagged' rule used for the binary
# distresses — and reproducible by sorting the compilation's ABSENTEE column
# largest-to-smallest and keeping the top row per FOLIO.
if 'ABSENTEE' in df.columns:
    ff_abs = (pd.to_numeric(df['ABSENTEE'], errors='coerce')
              .groupby(df['FOLIO'].astype(str).str.strip()).max())
else:
    ff_abs = pd.Series(dtype=float)

def absentee_counts_from_ff(folios):
    """3-way ABSENTEE counts (unique FOLIO) from fulfillment ff_abs, for C/D/E/F/G."""
    idx = pd.Series(list(folios)).dropna().astype(str).str.strip().unique()
    vals = ff_abs.reindex(idx)
    return {
        'Absentee':              int((vals == 1).sum()),
        'Absentee Out of State': int((vals == 2).sum()),
        'Owner Occupied':        int((vals == 0).sum()),
    }

C_dist   = {dtype: int(dist_long[dist_long['dtype'] == dtype]['FOLIO'].nunique())
            for dtype in distress_order}
C_likely = count_by_bin(ff_agg['ff_likely'], likely_bins, likely_labels)
C_score  = count_by_bin(ff_agg['ff_score'],  score_bins,  score_labels)
C_action = count_by_action(ff_agg['ff_days'], action_map)
# All fulfillment properties count in mkt — treat 0 contacts as 1 (recommended at least once)
ff_mkt_adj = ff_agg['ff_mkt'].fillna(0).clip(lower=1)
C_mkt    = count_by_bin(ff_mkt_adj, mkt_bins, mkt_labels)
print(f"  Column C ready: {len(ff_agg):,} properties")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: Columns F and G — fulfillment-matched leads/deals only
# ══════════════════════════════════════════════════════════════════════════════
_dl_files = sorted(
    f for f in DOCS_DIR.glob("*.xlsx")
    if re.search(r'leads|deals', f.name, re.IGNORECASE) and not f.name.startswith("~$")
)
if not _dl_files:
    raise FileNotFoundError(f"No leads/deals file found in {DOCS_DIR}")
print(f"\nLoading Deals and Leads ({len(_dl_files)} file(s)): {[f.name for f in _dl_files]}")
dl = pd.concat([pd.read_excel(f) for f in _dl_files], ignore_index=True)
dl['LIKELY DEAL SCORE']   = pd.to_numeric(dl['LIKELY DEAL SCORE'],   errors='coerce')
dl['SCORE']               = pd.to_numeric(dl['SCORE'],               errors='coerce')
for _col in ['MARKETING DM COUNT', 'MARKETING SMS COUNT']:
    dl[_col] = pd.to_numeric(dl[_col] if _col in dl.columns else 0, errors='coerce').fillna(0)
dl['dl_days'] = dl['ACTION PLANS'].apply(extract_days)
dl['dl_mkt']  = dl['MARKETING DM COUNT'] + dl['MARKETING SMS COUNT']
dl['FOLIO_key'] = dl['FOLIO'].astype(str).str.strip()
dl.loc[dl['FOLIO_key'] == 'nan', 'FOLIO_key'] = np.nan

# ── Filter leads/deals to on or after the START OF THE FULFILLMENT WINDOW ─────
# Attribution rule: a lead/deal can only be credited to this fulfillment if it
# occurred during the analysis window or after it (the latter captures the data-
# validation lag — deals that mature after WINDOW_END). Anything dated before the
# window predates the recommendation and cannot be its result, so it is excluded.
# This aligns F/G with Column D, which already uses SOLD_SINCE = WINDOW_START.
# (CLIENT_START_DATE is retained above for documentation; it is no longer the cutoff.)
LEADS_DEALS_SINCE = WINDOW_START
# Date filter per status:
#   Lead        → LEAD DATE        (blank date → included, can't rule it out)
#   Appointment → APPOINTMENT DATE (blank date → included; counted as Lead in report)
#   Dead Lead   → LEAD DATE        (blank date → included; counted as Lead in report)
#   Contract    → LEAD DATE (if available; blank date → included; counted as Lead in report)
#   Deal        → LAST SALE DATE   (must have a valid date)
print(f"  Rows before date filter: {len(dl):,}")
dl['_date_col'] = pd.NaT
dl.loc[dl['PROPERTY STATUS'] == 'Lead',        '_date_col'] = pd.to_datetime(dl.loc[dl['PROPERTY STATUS'] == 'Lead',        'LEAD DATE'],        errors='coerce')
dl.loc[dl['PROPERTY STATUS'] == 'Appointment', '_date_col'] = pd.to_datetime(dl.loc[dl['PROPERTY STATUS'] == 'Appointment', 'APPOINTMENT DATE'], errors='coerce')
dl.loc[dl['PROPERTY STATUS'] == 'Dead Lead',   '_date_col'] = pd.to_datetime(dl.loc[dl['PROPERTY STATUS'] == 'Dead Lead',   'LEAD DATE'],        errors='coerce')
dl.loc[dl['PROPERTY STATUS'] == 'Contract',    '_date_col'] = pd.to_datetime(dl.loc[dl['PROPERTY STATUS'] == 'Contract',    'LEAD DATE'],        errors='coerce')
dl.loc[dl['PROPERTY STATUS'] == 'Deal',        '_date_col'] = pd.to_datetime(dl.loc[dl['PROPERTY STATUS'] == 'Deal',        'LAST SALE DATE'],   errors='coerce')
# Contracts use LEAD DATE if available; blank → included as lead (uncertain, can't rule out)

# Lead-type statuses: include if date >= LEADS_DEALS_SINCE OR date is blank (uncertain, can't rule out)
# Deals: must have a valid LAST SALE DATE >= LEADS_DEALS_SINCE
_lead_types = dl['PROPERTY STATUS'].isin({'Lead', 'Appointment', 'Dead Lead', 'Contract'})
_keep = (
    (_lead_types & (dl['_date_col'].isna() | (dl['_date_col'] >= LEADS_DEALS_SINCE))) |
    (~_lead_types & (dl['_date_col'] >= LEADS_DEALS_SINCE))
)
dl = dl[_keep].drop(columns=['_date_col'])

# Appointments, Dead Leads, and Contracts counted as Leads in the final report
dl.loc[dl['PROPERTY STATUS'].isin(['Appointment', 'Dead Lead', 'Contract']), 'PROPERTY STATUS'] = 'Lead'

print(f"  Rows after date filter (>= {LEADS_DEALS_SINCE.date()} = window start, blanks kept for leads): {len(dl):,}")
print(f"    Leads (incl. appointments + dead leads): {(dl['PROPERTY STATUS'] == 'Lead').sum():,} | Deals: {(dl['PROPERTY STATUS'] == 'Deal').sum():,}")

ff_for_join = ff_agg.reset_index().rename(columns={'FOLIO': 'FOLIO_key'})
dl = dl.merge(ff_for_join, on='FOLIO_key', how='left')

BINARY_DIST_MAP = {
    'PRE-FORECLOSURE': 'Pre-foreclosure', 'VACANT': 'Vacant',
    'TAXES': 'Taxes (Tax Delinquent)', 'ESTATE': 'Estate (Pre-Probate)',
    'INTER FAMILY TRANSFER': 'Inter family transfer', 'PROBATE': 'Probate',
    'JUDGEMENT': 'Judgement', 'LIENS CITY/COUNTY': 'Liens city/county',
    'LIENS OTHER': 'Liens other', 'LIENS MECHANIC': 'Liens Mechanic',
    'HIGH EQUITY': 'High Equity', '55+': '55+ (Senior)',
    'ABSENTEE': 'Absentee', 'DOWNSIZING': 'Downsizing',
}

def get_binary_distress(row):
    types = set()
    for col, label in BINARY_DIST_MAP.items():
        if col in row.index:
            val = pd.to_numeric(row[col], errors='coerce')
            if not pd.isna(val) and val > 0:
                types.add(label)
    return types

def resolve_row(row):
    in_ff = pd.notna(row.get('ff_likely'))
    if in_ff:
        return pd.Series({
            'Data_Source': 'Fulfillment',
            'r_likely':    row['ff_likely'],
            'r_score':     row['ff_score'],
            'r_days':      row['ff_days'],
            'r_mkt':       max(row['ff_mkt'], 1) if pd.notna(row.get('ff_mkt')) else 1,
            'r_distress':  ff_distress_map.get(str(row['FOLIO_key']).strip(), set()),
        })
    else:
        return pd.Series({
            'Data_Source': 'Deals and Leads Atlas',
            'r_likely':    row['LIKELY DEAL SCORE'],
            'r_score':     row['SCORE'],
            'r_days':      row['dl_days'],
            'r_mkt':       row['dl_mkt'],
            'r_distress':  get_binary_distress(row),
        })

resolved = dl.apply(resolve_row, axis=1)
dl = pd.concat([dl, resolved], axis=1)

def compute_col_counts(sub_df):
    likely_d = count_by_bin(sub_df['r_likely'], likely_bins, likely_labels)
    score_d  = count_by_bin(sub_df['r_score'],  score_bins,  score_labels)
    action_d = count_by_action(sub_df['r_days'], action_map)
    mkt_d    = count_by_bin(sub_df['r_mkt'],    mkt_bins,   mkt_labels)
    dist_d   = {dtype: int(
        sub_df['r_distress'].apply(lambda s: dtype in s if isinstance(s, set) else False).sum()
    ) for dtype in distress_order}
    return likely_d, score_d, action_d, mkt_d, dist_d

leads_ff = dl[(dl['PROPERTY STATUS'] == 'Lead') & (dl['Data_Source'] == 'Fulfillment')].copy()
deals_ff = dl[(dl['PROPERTY STATUS'] == 'Deal') & (dl['Data_Source'] == 'Fulfillment')].copy()
F_likely, F_score, F_action, F_mkt, F_dist = compute_col_counts(leads_ff)
G_likely, G_score, G_action, G_mkt, G_dist = compute_col_counts(deals_ff)
print(f"  Columns F/G ready — Leads: {len(leads_ff)} | Deals: {len(deals_ff)}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3: Load domain → Column B + prep for D and E
# ══════════════════════════════════════════════════════════════════════════════
print("\nLoading domain files...")
_parquet = DOMAIN_DIR / "domain.parquet"
import os as _os
if _os.path.exists(_parquet):
    # Strip pandas metadata before to_pandas() so StringDtype columns load as
    # plain object — pandas StringDtype reconstruction balloons memory on this dataset.
    import pyarrow.parquet as _pq
    _table = _pq.read_table(_parquet).replace_schema_metadata(None)
    dom = _table.to_pandas()
    del _table
    print(f"  Loaded from parquet: {len(dom):,} rows")
else:
    print("  Parquet not found — reading xlsx and converting (first run is slower)...")
    _xlsx = sorted(f for f in DOMAIN_DIR.glob("*.xlsx") if not f.name.startswith("~$"))
    dom   = pd.concat([pd.read_excel(f) for f in _xlsx], ignore_index=True)
    for _c in dom.select_dtypes(include='object').columns:
        dom[_c] = dom[_c].astype(str)
    dom.to_parquet(_parquet, index=False)
    print(f"  Saved parquet cache: {_parquet.name}")
print(f"  {len(dom):,} rows, {dom['FOLIO'].nunique():,} unique FOLIOs")

dom['LIKELY DEAL SCORE']   = pd.to_numeric(dom['LIKELY DEAL SCORE'],   errors='coerce')
dom['SCORE']               = pd.to_numeric(dom['SCORE'],               errors='coerce')
for _col in ['MARKETING DM COUNT', 'MARKETING SMS COUNT']:
    dom[_col] = pd.to_numeric(dom[_col] if _col in dom.columns else 0, errors='coerce').fillna(0)
dom['total_mkt']           = dom[['MARKETING DM COUNT', 'MARKETING SMS COUNT']].max(axis=1)
dom['ACTION_DAYS']         = dom['ACTION PLANS'].apply(extract_days)
dom['LAST SALE DATE']      = pd.to_datetime(dom['LAST SALE DATE'], errors='coerce')
dom['buybox_int']          = pd.to_numeric(dom['PROPERTY ID (BUYBOX)'], errors='coerce')

# Dedupe by FOLIO — domain has ~167K duplicate FOLIO rows that would otherwise
# inflate Column B / D / E counts. Keep the row with the most recent LAST SALE DATE
# so newer sale info wins; NaT dates sort first so they're dropped when a real date exists.
_pre_dedupe = len(dom)
dom = (dom.sort_values('LAST SALE DATE', na_position='first')
          .drop_duplicates('FOLIO', keep='last')
          .reset_index(drop=True))
print(f"  After dedupe by FOLIO: {len(dom):,} rows ({_pre_dedupe - len(dom):,} dropped)")

# ── ABSENTEE (3-way) for Columns C / F / G — from FULFILLMENT ABSENTEE ────────
# ABSENTEE is a 3-way classification (0 = owner-occupied, 1 = in-state absentee,
# 2 = out-of-state), not a binary flag, so it is excluded from DOMAIN_DIST_MAP /
# dist_long / ff_distress_map. C (built from dist_long) and F/G (built from
# ff_distress_map) therefore carry NO absentee signal on their own — patch all
# three classes here from the fulfillment ff_abs map (max across months), the
# same source D/E use. Domain is NOT used: it holds only current ownership, not
# the point-in-time status when the property was recommended.
print("\nPatching Absentee / Owner Occupied for Columns C/F/G from fulfillment ABSENTEE (max across months)...")
C_dist.update(absentee_counts_from_ff(ff_agg.index))
F_dist.update(absentee_counts_from_ff(leads_ff['FOLIO_key']))
G_dist.update(absentee_counts_from_ff(deals_ff['FOLIO_key']))
print(f"  Absentee              — C: {C_dist['Absentee']:,} | F: {F_dist['Absentee']:,} | G: {G_dist['Absentee']:,}")
print(f"  Absentee Out of State — C: {C_dist['Absentee Out of State']:,} | F: {F_dist['Absentee Out of State']:,} | G: {G_dist['Absentee Out of State']:,}")
print(f"  Owner Occupied        — C: {C_dist['Owner Occupied']:,} | F: {F_dist['Owner Occupied']:,} | G: {G_dist['Owner Occupied']:,}")

# Column B — filter to BUYBOX SCORE > 0
dom['BUYBOX SCORE'] = pd.to_numeric(dom['BUYBOX SCORE'], errors='coerce').fillna(0)
dom_b = dom[dom['BUYBOX SCORE'] > 0].copy()
print(f"  Domain with BUYBOX SCORE > 0: {len(dom_b):,}")

B_likely = count_by_bin(dom_b['LIKELY DEAL SCORE'], likely_bins, likely_labels)
B_score  = count_by_bin(dom_b['SCORE'],             score_bins,  score_labels)
B_action = count_by_action(dom_b['ACTION_DAYS'],    action_map)
B_mkt    = count_by_bin(dom_b[dom_b['total_mkt'] > 0]['total_mkt'], mkt_bins, mkt_labels)
B_dist   = domain_distress_counts(dom_b)
print(f"  Column B ready — total: {sum(B_likely.values()):,}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4: Column D — Sold since start of fulfillment window
# ══════════════════════════════════════════════════════════════════════════════
print(f"\nComputing Column D (sold since {SOLD_SINCE.strftime('%Y-%m-%d')}, fulfillment only)...")
sold_all = dom[dom['LAST SALE DATE'] >= SOLD_SINCE].copy()
ff_join  = ff_agg[['ff_likely', 'ff_score', 'ff_days', 'ff_mkt']].reset_index()
sold_all = sold_all.merge(ff_join, on='FOLIO', how='left')
# Restrict to fulfillment-recommended properties only
sold = sold_all[sold_all['ff_likely'].notna()].copy()
print(f"  Sold since {SOLD_SINCE.strftime('%Y-%m-%d')}: {len(sold_all):,} total | In fulfillment: {len(sold):,}")

sold['d_likely'] = sold['ff_likely']
sold['d_score']  = sold['ff_score']
sold['d_days']   = sold['ff_days']
sold['d_mkt']    = sold['ff_mkt'].fillna(0).clip(lower=1)

D_likely = count_by_bin(sold['d_likely'], likely_bins, likely_labels)
D_score  = count_by_bin(sold['d_score'],  score_bins,  score_labels)
D_action = count_by_action(sold['d_days'], action_map)
D_mkt    = count_by_bin(sold['d_mkt'], mkt_bins, mkt_labels)
D_dist   = fulfillment_distress_counts(sold, dist_long)
print(f"  Column D ready — total: {sum(D_likely.values()):,}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5: Column E — Market Deals
# Properties believed to have been bought at a significant discount — typically bought and resold
# for a profit in a short period of time. Computed as the overlap between the Market Deals file
# and the fulfillment-sold subset (Column D). Same score resolution as Column D: pre-sale fulfillment MAX values.
# ══════════════════════════════════════════════════════════════════════════════
print("\nComputing Column E (market deals x fulfillment)...")
md = pd.read_excel(MARKET_PATH)

if 'PropertyID' in md.columns:
    md_ids = set(md['PropertyID'].dropna().astype(int).unique())
    ff_sold_buybox_ids = set(sold['buybox_int'].dropna().astype(int).unique())
    overlap_ids = md_ids & ff_sold_buybox_ids
    print(f"  Match by PropertyID. Market deals: {len(md_ids):,} | Fulfillment-sold: {len(ff_sold_buybox_ids):,} | Overlap: {len(overlap_ids):,}")
    mkt_sold = sold[sold['buybox_int'].isin(overlap_ids)].copy()
else:
    # Fallback: market deals file has no PropertyID — match on normalized ADDRESS + ZIP.
    _SUFFIX_MAP = {
        r'\bSTREET\b': 'ST',  r'\bAVENUE\b': 'AVE', r'\bROAD\b': 'RD',
        r'\bBOULEVARD\b': 'BLVD', r'\bDRIVE\b': 'DR', r'\bLANE\b': 'LN',
        r'\bCOURT\b': 'CT', r'\bCIRCLE\b': 'CIR', r'\bTRAIL\b': 'TRL',
        r'\bPLACE\b': 'PL', r'\bPARKWAY\b': 'PKWY', r'\bHIGHWAY\b': 'HWY',
        r'\bTERRACE\b': 'TER',
    }
    def _norm_addr(s):
        if pd.isna(s): return None
        s = re.sub(r'[.,#]', '', str(s).upper().strip())
        for pat, rep in _SUFFIX_MAP.items():
            s = re.sub(pat, rep, s)
        return re.sub(r'\s+', ' ', s) or None
    def _norm_zip(z):
        if pd.isna(z): return None
        s = re.sub(r'\D', '', str(z))[:5]
        return s if len(s) == 5 else None
    def _key(a, z):
        na, nz = _norm_addr(a), _norm_zip(z)
        return f"{na}|{nz}" if na and nz else None

    md_cols = {c.lower(): c for c in md.columns}
    addr_col = md_cols.get('street address') or md_cols.get('address')
    zip_col  = md_cols.get('zip code')       or md_cols.get('zip')
    if not addr_col or not zip_col:
        raise KeyError(f"Market deals missing PropertyID and address/zip. Got: {list(md.columns)}")

    md['_addr_key']   = [_key(a, z) for a, z in zip(md[addr_col], md[zip_col])]
    sold['_addr_key'] = [_key(a, z) for a, z in zip(sold['ADDRESS'], sold['ZIP'])]
    md_keys     = set(md['_addr_key'].dropna().unique())
    sold_keys   = set(sold['_addr_key'].dropna().unique())
    overlap_keys = md_keys & sold_keys
    print(f"  Match by ADDRESS+ZIP (no PropertyID in market deals). "
          f"Market deals keys: {len(md_keys):,} | Fulfillment-sold keys: {len(sold_keys):,} | Overlap: {len(overlap_keys):,}")
    mkt_sold = sold[sold['_addr_key'].isin(overlap_keys)].copy()

E_likely = count_by_bin(mkt_sold['d_likely'], likely_bins, likely_labels)
E_score  = count_by_bin(mkt_sold['d_score'],  score_bins,  score_labels)
E_action = count_by_action(mkt_sold['d_days'], action_map)
E_mkt    = count_by_bin(mkt_sold['d_mkt'], mkt_bins, mkt_labels)
E_dist   = fulfillment_distress_counts(mkt_sold, dist_long)
print(f"  Column E ready — total: {sum(E_likely.values()):,}")

print(f"\n=== COLUMN E (market deals) ===")
print("Likely Deal Score:", E_likely)
print("Action Plan:", E_action)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5b: Enrich fulfillment compilation for validation, then save
# Adds columns the user can use to spot-check D, F/G, and E:
#   • LAST SALE DATE   — most recent sale per FOLIO from domain (validates D)
#   • PROPERTY STATUS  — Lead/Deal from deals/leads file (validates F/G)
#   • MARKET DEAL      — Yes/No flag from market deals overlap (validates E)
#   • CLIENT LEAD      — Yes/No flag, FOLIO is in the report's Column F population
#   • CLIENT DEAL      — Yes/No flag, FOLIO is in the report's Column G population
# ══════════════════════════════════════════════════════════════════════════════
print("\nEnriching fulfillment compilation with validation columns...")

_last_sale = (dom[['FOLIO', 'LAST SALE DATE']]
              .assign(FOLIO=dom['FOLIO'].astype(str).str.strip())
              .dropna(subset=['LAST SALE DATE'])
              .sort_values('LAST SALE DATE')
              .drop_duplicates('FOLIO', keep='last')
              .set_index('FOLIO')['LAST SALE DATE'])

# Prefer Deal over Lead when a FOLIO appears as both
_status_priority = {'Deal': 2, 'Lead': 1}
_dl_status = (dl[['FOLIO_key', 'PROPERTY STATUS']]
              .dropna(subset=['FOLIO_key'])
              .assign(_p=lambda x: x['PROPERTY STATUS'].map(_status_priority).fillna(0))
              .sort_values('_p')
              .drop_duplicates('FOLIO_key', keep='last')
              .set_index('FOLIO_key')['PROPERTY STATUS'])

_mkt_folios = set(mkt_sold['FOLIO'].astype(str).str.strip().unique())

# Client Lead / Client Deal flags — derived from the SAME populations the report
# uses for Columns F and G (leads_ff / deals_ff), so a "Yes" count reconciles
# against the report's unique-FOLIO membership for those columns. Note: F/G in the
# report count rows, so a FOLIO with multiple leads contributes 1 "Yes" here but
# may contribute >1 to the report total.
_lead_folios = set(leads_ff['FOLIO_key'].dropna().astype(str).str.strip().unique())
_deal_folios = set(deals_ff['FOLIO_key'].dropna().astype(str).str.strip().unique())

_df_folio_key = df['FOLIO'].astype(str).str.strip()
df['LAST SALE DATE']  = _df_folio_key.map(_last_sale)
df['PROPERTY STATUS'] = _df_folio_key.map(_dl_status).fillna('')
df['MARKET DEAL']     = np.where(_df_folio_key.isin(_mkt_folios), 'Yes', 'No')
df['CLIENT LEAD']     = np.where(_df_folio_key.isin(_lead_folios), 'Yes', 'No')
df['CLIENT DEAL']     = np.where(_df_folio_key.isin(_deal_folios), 'Yes', 'No')
# Counts below are UNIQUE FOLIOs (= report F/G membership). The compilation has
# ~one row per FOLIO-month, so a filter on CLIENT LEAD=Yes returns more *rows*
# than this — each flagged FOLIO repeats across the months it was recommended.
print(f"  Audit flags (unique FOLIOs) — "
      f"MARKET DEAL: {len(_mkt_folios):,} | "
      f"CLIENT LEAD: {len(_lead_folios):,} (= report Col F) | "
      f"CLIENT DEAL: {len(_deal_folios):,} (= report Col G)")

print(f"  Saving Fulfillment_Compilation.xlsx with validation columns...")
# strings_to_urls=False  — disables xlsxwriter's URL auto-detection
#   (fulfillment data contains >65,530 property URLs, exceeding Excel's per-sheet
#   limit and causing a warning flood + memory blowup).
# tmpdir=BASE  — keeps xlsxwriter's intermediate files on the project drive;
#   the system %TEMP% is on C: which can be full on this machine.
with pd.ExcelWriter(COMPILED_PATH, engine='xlsxwriter',
                    engine_kwargs={'options': {'strings_to_urls': False,
                                                'tmpdir': str(BASE)}}) as _writer:
    df.to_excel(_writer, index=False)
print(f"  Saved: {COMPILED_PATH.name}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5c: ZIP-level breakdown block
# One row per ZIP across the same six populations (B–G), so the client can see how
# each of their zips is behaving. Ordered by Client Deals (Column G) descending —
# "the zips with the most deals first" — with Sold (D) then Fulfillment (C) as
# tiebreakers so the zero-deal tail still sorts sensibly.
#
# Universe excludes BUYBOX SCORE 0: the domain side uses dom_b (already > 0), and
# every fulfillment-recommended property is > 0, so C/D/E/F/G inherit the rule and
# the ZIP-block column totals reconcile with the main report's column totals.
# ZIP set = union of domain (buybox>0) zips and fulfillment zips.
# ══════════════════════════════════════════════════════════════════════════════
print("\nComputing ZIP-level breakdown block...")

def _norm_zip(z):
    if pd.isna(z): return None
    s = re.sub(r'\D', '', str(z))[:5]
    return s if len(s) == 5 else None

# FOLIO → ZIP from the fulfillment compilation (zip is stable across months for a
# FOLIO; keep the first valid one). Used to attribute Column C FOLIOs to a zip.
_dfz        = df[['FOLIO', 'ZIP']].copy()
_dfz['_z']  = _dfz['ZIP'].map(_norm_zip)
_dfz['_fk'] = _dfz['FOLIO'].astype(str).str.strip()
_ff_zipmap  = _dfz.dropna(subset=['_z']).drop_duplicates('_fk').set_index('_fk')['_z']

# Per-population ZIP counts (same populations the main report's columns use)
B_zip = dom_b['ZIP'].map(_norm_zip).value_counts()                                  # Total Properties
C_zip = pd.Series(ff_agg.index.astype(str).str.strip()).map(_ff_zipmap).value_counts()  # In fulfillment
D_zip = sold['ZIP'].map(_norm_zip).value_counts()                                   # Sold
E_zip = mkt_sold['ZIP'].map(_norm_zip).value_counts()                               # Market Deals
F_zip = leads_ff['ZIP'].map(_norm_zip).value_counts()                               # Client Leads (rows)
G_zip = deals_ff['ZIP'].map(_norm_zip).value_counts()                               # Client Deals (rows)

_all_zips = (set(B_zip.index) | set(C_zip.index) | set(D_zip.index)
             | set(E_zip.index) | set(F_zip.index) | set(G_zip.index))
_all_zips.discard(None)

zip_table = [
    (z,
     int(B_zip.get(z, 0)), int(C_zip.get(z, 0)), int(D_zip.get(z, 0)),
     int(E_zip.get(z, 0)), int(F_zip.get(z, 0)), int(G_zip.get(z, 0)))
    for z in _all_zips
]
# Sort: Client Deals (G) desc, then Sold (D) desc, then Fulfillment (C) desc, then zip asc
zip_table.sort(key=lambda r: (-r[6], -r[3], -r[2], r[0]))
print(f"  ZIP block: {len(zip_table):,} zips "
      f"({sum(1 for r in zip_table if r[6] > 0):,} with >=1 client deal)")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6: Build report rows
# ══════════════════════════════════════════════════════════════════════════════
def build_rows(B_lik, B_sco, B_act, B_mkt, B_dis,
               C_lik, C_sco, C_act, C_mkt, C_dis,
               D_lik, D_sco, D_act, D_mkt, D_dis,
               E_lik, E_sco, E_act, E_mkt, E_dis,
               F_lik, F_sco, F_act, F_mkt, F_dis,
               G_lik, G_sco, G_act, G_mkt, G_dis,
               zip_rows=None):
    n = len(COLS)

    def section_header(name):
        r = list(COLS); r[0] = name; return r

    def data_row(label, b, c, d, e, f, g):
        r = [''] * n
        r[0] = label
        r[1] = b    # B — Total Properties
        r[2] = c    # C — Properties in the fulfillment
        r[3] = d    # D — Sold
        r[4] = e    # E — Market Deals
        r[5] = f    # F — Clients Lead
        r[6] = g    # G — Client Deals
        # H–K left empty here; formulas written during Excel rendering
        return r

    def blank():
        return [''] * n

    rows = []

    rows.append(section_header('Likely Deal Score'))
    for lbl in ['100-81', '80-61', '60-41', '40-21', '20-0']:
        rows.append(data_row(lbl, B_lik[lbl], C_lik[lbl], D_lik[lbl], E_lik[lbl], F_lik[lbl], G_lik[lbl]))
    rows.append(data_row('Total', sum(B_lik.values()), sum(C_lik.values()), sum(D_lik.values()), sum(E_lik.values()), sum(F_lik.values()), sum(G_lik.values())))
    rows.append(blank())

    rows.append(section_header('Total Score'))
    for lbl in ['1000-801', '800-601', '600-401', '400-201', '200-0']:
        rows.append(data_row(lbl, B_sco[lbl], C_sco[lbl], D_sco[lbl], E_sco[lbl], F_sco[lbl], G_sco[lbl]))
    rows.append(data_row('Total', sum(B_sco.values()), sum(C_sco.values()), sum(D_sco.values()), sum(E_sco.values()), sum(F_sco.values()), sum(G_sco.values())))
    rows.append(blank())

    rows.append(section_header('Action Plan'))
    for lbl in ['30 day', '60 day', '90 day']:
        rows.append(data_row(lbl, B_act[lbl], C_act[lbl], D_act[lbl], E_act[lbl], F_act[lbl], G_act[lbl]))
    rows.append(data_row('Total', sum(B_act.values()), sum(C_act.values()), sum(D_act.values()), sum(E_act.values()), sum(F_act.values()), sum(G_act.values())))
    rows.append(blank())

    rows.append(section_header('Mkt Count'))
    for lbl in ['1 to 5', '6 to 10', '11 to 19', '20+']:
        rows.append(data_row(lbl, B_mkt[lbl], C_mkt[lbl], D_mkt[lbl], E_mkt[lbl], F_mkt[lbl], G_mkt[lbl]))
    rows.append(data_row('Total', sum(B_mkt.values()), sum(C_mkt.values()), sum(D_mkt.values()), sum(E_mkt.values()), sum(F_mkt.values()), sum(G_mkt.values())))
    rows.append(blank())

    rows.append(section_header('Distress'))
    for dtype in distress_order:
        rows.append(data_row(dtype, B_dis[dtype], C_dis[dtype], D_dis[dtype], E_dis[dtype], F_dis[dtype], G_dis[dtype]))
    rows.append(data_row('Total', sum(B_dis.values()), sum(C_dis.values()), sum(D_dis.values()), sum(E_dis.values()), sum(F_dis.values()), sum(G_dis.values())))

    # ── ZIP Code block — one row per zip, ordered by Client Deals desc ──────────
    if zip_rows:
        rows.append(blank())
        rows.append(section_header(ZIP_SECTION_TITLE))
        tB = tC = tD = tE = tF = tG = 0
        for z, b, c, d, e, f, g in zip_rows:
            rows.append(data_row(z, b, c, d, e, f, g))
            tB += b; tC += c; tD += d; tE += e; tF += f; tG += g
        rows.append(data_row('Total', tB, tC, tD, tE, tF, tG))

    return rows


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7: Write Excel with formulas for H–K
# ══════════════════════════════════════════════════════════════════════════════
def write_excel(path, rows, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Styles
    header_font   = Font(name='Helvetica', bold=True, size=10, color='FFFFFF')
    header_fill   = PatternFill('solid', fgColor='0B5394')
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    section_font  = Font(name='Georgia', bold=True, size=10)
    section_fill  = PatternFill('solid', fgColor='D9E2F3')
    section_align = Alignment(horizontal='left', vertical='center')
    total_font    = Font(name='Helvetica', bold=True, size=10)
    total_fill    = PatternFill('solid', fgColor='F2F2F2')
    data_font     = Font(name='Helvetica', size=10)
    count_font    = Font(name='Helvetica', bold=True, size=10, color='0B5394')
    pct_font      = Font(name='Helvetica', size=10, color='444444')
    pct_font_bold = Font(name='Helvetica', bold=True, size=10, color='444444')
    center_align  = Alignment(horizontal='center', vertical='center')
    left_align    = Alignment(horizontal='left',   vertical='center')
    thin_side     = Side(style='thin', color='CCCCCC')
    thin          = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

    # Header row
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = thin
    ws.row_dimensions[1].height = 36

    # Header comments — one per column, visible as hover tooltips in Excel
    header_comments = {
        'A': (
            "Category\n"
            "Row labels for each section. Section headers: Likely Deal Score, Total Score, "
            "Action Plan, Mkt Count, Distress. Sub-rows are the range buckets or distress "
            "types within each section, plus a Total row at the bottom of each section."
        ),
        'B': (
            "Total Properties\n"
            "All properties in the domain with BUYBOX SCORE > 0 — the full addressable universe.\n"
            "Likely Deal Score / Total Score: all scored properties.\n"
            "Action Plan: only properties that have an action plan assigned.\n"
            "Mkt Count: only properties with at least 1 marketing contact (DM or SMS).\n"
            "Distress: derived from binary domain columns; DEFAULT RISK is a direct domain column.\n"
            "Owner Occupied: ABSENTEE == 0 (NaN excluded — unknown is not owner-occupied)."
        ),
        'C': (
            "Properties in the Fulfillment\n"
            "Unique properties (by FOLIO) recommended across all fulfillment files in the "
            "3-month analysis window.\n"
            "Scores: MAX across all months per property.\n"
            "Action Plan: best (lowest day count) across all months.\n"
            "Mkt Count: max(total DM contacts, total SMS contacts) — strongest channel only; "
            "properties with 0 contacts are treated as 1 (recommended at least once).\n"
            "Distress: fulfillment binary distress columns (PRE-FORECLOSURE > 0, "
            "VACANT > 0, TAXES > 0, …) per FOLIO across all months + Default Risk "
            "cross-referenced from domain.\n"
            "Absentee / Absentee Out of State / Owner Occupied: fulfillment ABSENTEE "
            "code (0/1/2), max across months per FOLIO."
        ),
        'D': (
            f"Sold\n"
            f"Properties sold since {SOLD_SINCE.strftime('%b %d %Y')} that were also recommended in a fulfillment.\n"
            f"Cutoff = start of the {WINDOW_START.strftime('%Y-%m')} – {WINDOW_END.strftime('%Y-%m')} fulfillment window.\n"
            f"Only fulfillment-matched properties are included (domain-only sold properties excluded).\n"
            f"Scores are pre-sale fulfillment MAX values (snapshot before the sale).\n"
            f"Distress: fulfillment binary distress columns per FOLIO (same source as C); "
            f"ABSENTEE (Absentee / Out of State / Owner Occupied) = fulfillment ABSENTEE, max across months."
        ),
        'E': (
            "Market Deals\n"
            "Properties believed to have been bought at a significant discount — typically bought and resold "
            "for a profit in a short period of time. These represent investor market transactions that "
            "overlapped with properties the client was recommended.\n"
            "Source: properties that appear in both the Market Deals file and the fulfillment-sold subset.\n"
            "Join: Market Deals PropertyID matched to domain PROPERTY ID (BUYBOX).\n"
            "Scores: pre-sale fulfillment MAX values (same as Column D).\n"
            "Distress: fulfillment binary distress columns per FOLIO (same source as C/D); "
            "ABSENTEE (Absentee / Out of State / Owner Occupied) = fulfillment ABSENTEE, max across months."
        ),
        'F': (
            "Clients Lead\n"
            "Leads, Appointments, Dead Leads, and Contracts from the Deals & Leads file "
            "that matched a fulfillment property via FOLIO.\n"
            "Status mapping: Lead, Appointment, Dead Lead, and Contract all count as a Lead here.\n"
            "Date filter: Lead Date or Appointment Date >= client start date. "
            "Blank dates are included — they cannot be ruled out.\n"
            "Contracts use Lead Date if available; blank date — always included.\n"
            "Only fulfillment-matched rows are counted. Unmatched rows are excluded.\n"
            "Scores: fulfillment MAX values. Mkt: 0 contacts treated as 1."
        ),
        'G': (
            "Client Deals\n"
            "Deals from the Deals & Leads file that matched a fulfillment property via FOLIO.\n"
            "Date filter: Last Sale Date >= client start date (must have a valid date).\n"
            "Only fulfillment-matched deals are counted. Unmatched deals are excluded.\n"
            "Scores: fulfillment MAX values."
        ),
        'H': (
            "Sold Concentration %\n"
            "Formula: =D/C\n"
            f"Share of fulfillment-recommended properties that were sold since {SOLD_SINCE.strftime('%b %d %Y')}. "
            "Measures how many of the recommended properties transacted."
        ),
        'I': (
            "Sold to Investors Concentration\n"
            "Formula: =E/C\n"
            "Share of fulfillment-recommended properties that were sold to investors (market deals). "
            "Measures investor acquisition rate within the fulfillment population."
        ),
        'J': (
            "Client Deals Concentration\n"
            "Formula: =G/C\n"
            "Share of fulfillment-recommended properties that resulted in a client deal. "
            "Measures deal conversion rate within the fulfillment population."
        ),
        'K': (
            "Client Leads Concentration\n"
            "Formula: =F/C\n"
            "Share of fulfillment-recommended properties that generated a client lead "
            "(includes Leads, Appointments, Dead Leads, and Contracts). "
            "Measures lead engagement rate within the fulfillment population."
        ),
    }
    for col_letter, text in header_comments.items():
        cell = ws[f'{col_letter}1']
        cmt = Comment(text, "Atlas Report")
        cmt.width  = 400
        cmt.height = 200
        cell.comment = cmt


    # Count columns (0-based): B=1, C=2, D=3, E(Market Deals)=4, F(Leads)=5, G(Deals)=6
    COUNT_COLS = [1, 2, 3, 4, 5, 6]
    # Formula columns (0-based): H=7, I=8, J=9, K=10
    # Formulas reference 1-based Excel columns:
    #   H = D/C  →  =IFERROR(D{r}/C{r},"")   Sold Concentration
    #   I = E/C  →  =IFERROR(E{r}/C{r},"")   Sold to Investors Concentration
    #   J = G/C  →  =IFERROR(G{r}/C{r},"")   Client Deals Concentration
    #   K = F/C  →  =IFERROR(F{r}/C{r},"")   Client Leads Concentration

    for row_data in rows:
        ws.append(row_data)
        r = ws.max_row
        label = row_data[0]

        if label in SECTION_NAMES:
            for cell in ws[r]:
                cell.font = section_font; cell.fill = section_fill
                cell.alignment = section_align; cell.border = thin

        elif label == '':
            pass  # blank separator

        else:
            # Data row or Total row
            is_total = (label == 'Total')
            base_font  = total_font if is_total else data_font
            base_fill  = total_fill if is_total else None
            c_font     = Font(name='Helvetica', bold=True, size=10, color='0B5394') if is_total else count_font
            p_font     = pct_font_bold if is_total else pct_font

            for cell in ws[r]:
                cell.font      = base_font
                cell.alignment = left_align
                cell.border    = thin
                if base_fill:
                    cell.fill = base_fill

            # Style count cells B–G
            for ci in COUNT_COLS:
                ws[r][ci].font      = c_font
                ws[r][ci].alignment = center_align

            # Write formulas for H–K and format as percentage
            ws[r][7].value  = f'=IFERROR(D{r}/C{r},"")'   # H: Sold / Fulfillment
            ws[r][8].value  = f'=IFERROR(E{r}/C{r},"")'   # I: Market Deals / Fulfillment
            ws[r][9].value  = f'=IFERROR(G{r}/C{r},"")'   # J: Client Deals / Fulfillment
            ws[r][10].value = f'=IFERROR(F{r}/C{r},"")'   # K: Client Leads / Fulfillment

            pct_fmt = '0.0%'
            for ci in [7, 8, 9, 10]:
                ws[r][ci].number_format = pct_fmt
                ws[r][ci].font          = p_font
                ws[r][ci].alignment     = center_align
                ws[r][ci].border        = thin

    # ── Highlight highs/lows per column within each section (green = high /
    #    more opportunity, red = low / less), so the client can scan it. Totals
    #    excluded. Distress and ZIP use top-3 / bottom-3; every other section
    #    uses just the single top & bottom. k is capped at rows//2 so the top
    #    and bottom sets never overlap; columns where all values are equal are
    #    left plain.
    hi_fill = PatternFill('solid', fgColor='C6EFCE')   # light green
    hi_font = Font(name='Helvetica', bold=True, size=10, color='006100')
    lo_fill = PatternFill('solid', fgColor='FFC7CE')   # light red
    lo_font = Font(name='Helvetica', bold=True, size=10, color='9C0006')
    THREE_EACH = {'Distress', ZIP_SECTION_TITLE}       # rest get top/bottom 1

    sections, current = [], None
    for i, row_data in enumerate(rows):
        ws_row, label = i + 2, row_data[0]
        if label in SECTION_NAMES:
            current = {'name': label, 'rows': []}
            sections.append(current)
        elif label == '':
            current = None
        elif label == 'Total':
            continue                      # exclude section totals from highs/lows
        elif current is not None:
            current['rows'].append((ws_row, row_data))

    for sec in sections:
        rws = sec['rows']
        cap = 3 if sec['name'] in THREE_EACH else 1
        k = min(cap, len(rws) // 2)
        if k == 0:
            continue
        for ci in COUNT_COLS:             # 0-based row_data index; Excel col = ci+1
            vals = [(wr, rd[ci]) for wr, rd in rws if isinstance(rd[ci], (int, float))]
            if len(vals) < 2 or max(v for _, v in vals) == min(v for _, v in vals):
                continue
            ranked   = sorted(vals, key=lambda t: t[1], reverse=True)
            top_rows = {wr for wr, _ in ranked[:k]}
            low_rows = {wr for wr, _ in ranked[-k:]} - top_rows
            for wr in top_rows:
                c = ws.cell(row=wr, column=ci + 1); c.fill = hi_fill; c.font = hi_font
            for wr in low_rows:
                c = ws.cell(row=wr, column=ci + 1); c.fill = lo_fill; c.font = lo_font

    # Legend for the highlighting (off to the right of the table)
    ws.cell(row=1, column=13, value='Highest per section (most)').fill = hi_fill
    ws.cell(row=1, column=13).font = hi_font
    ws.cell(row=2, column=13, value='Lowest per section (least)').fill = lo_fill
    ws.cell(row=2, column=13).font = lo_font
    ws.column_dimensions['M'].width = 26

    # Column widths
    ws.column_dimensions['A'].width = 32
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 22
    ws.freeze_panes = 'A2'

    wb.save(path)
    print(f"  Saved: {path}")


rows = build_rows(
    B_likely, B_score, B_action, B_mkt, B_dist,
    C_likely, C_score, C_action, C_mkt, C_dist,
    D_likely, D_score, D_action, D_mkt, D_dist,
    E_likely, E_score, E_action, E_mkt, E_dist,
    F_likely, F_score, F_action, F_mkt, F_dist,
    G_likely, G_score, G_action, G_mkt, G_dist,
    zip_rows=zip_table,
)

print(f"\nWriting report...")
write_excel(OUTPUT, rows, "Max Logic - BCDEFG")
print("\nAll done.")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 6: Companion Fulfillment Distress Analysis (Atlas-style HTML + PDF)
# Generated automatically from the SAME data, on the SAME window, so the
# headline "sold properties" count reconciles with Column D above. Wrapped in
# try/except so a template/Playwright issue never blocks the Excel report.
# ══════════════════════════════════════════════════════════════════════════════
try:
    from build_historical_report import generate_historical
    generate_historical(
        client_name=CLIENT_NAME,
        window_start=WINDOW_START,
        window_end=WINDOW_END,
        ff=df,
        dom=dom,
        out_dir=BASE,
        # Pass the Excel's per-FOLIO distress so the companion's sold-property
        # breakdown reconciles exactly with Column D (same source as C/D/E/F/G).
        ff_distress_map=ff_distress_map,
        ff_abs=ff_abs,
        dist_map=DOMAIN_DIST_MAP,
    )
except Exception as _e:
    print(f"\n[WARN] Companion HTML/PDF report not generated: {_e}")
    print("       The Excel distress report above is unaffected.")
